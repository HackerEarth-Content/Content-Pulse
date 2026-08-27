"""XLSX and CSV builders.

Every cell of user-entered text goes through `safe()`. Notes and customer names
are free text that lands in a spreadsheet, and Excel executes any cell starting
with = + - @ — the Django exports wrote those raw.
"""

from __future__ import annotations

import csv
import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from core.dates import today
from core.orm import STATUSES
from services import content_requests as cr_svc
from services import entries as entries_svc

MAX_ROWS = 20_000

HEAD_FILL = PatternFill("solid", fgColor="0F1620")
HEAD_FONT = Font(name="Calibri", bold=True, color="4FFFB0", size=10)
NAME_FILL = PatternFill("solid", fgColor="162032")
NAME_FONT = Font(name="Calibri", bold=True, color="9ECFFF", size=10)
DATE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
STRIPE = PatternFill("solid", fgColor="0C1117")
RULE = Border(bottom=Side(style="thin", color="1A2535"))

STATUS_LABEL = {
    "open": "Open",
    "in_progress": "In Progress",
    "blocked": "Blocked",
    "closed": "Done",
}

# One accent per sheet, cycled in order — the same palette the app itself
# uses for section headings (theme.css --accent-*), so a downloaded workbook
# reads as an extension of the screen rather than a generic export.
ACCENTS = ["3987E5", "199E70", "9085E9", "D55181", "D95926", "C98500", "E66767"]

_RISKY = re.compile(r"^[=+\-@\t\r]")


def safe(value):
    """Neutralise spreadsheet formula injection. A note like
    `=HYPERLINK("http://evil","click")` is a live formula in Excel, not text."""
    if isinstance(value, str) and _RISKY.match(value):
        return "'" + value
    return value


def _mix(hex_a: str, hex_b: str, t: float) -> str:
    """Blend hex_a toward hex_b by fraction t — used to tint the zebra stripe
    with a sheet's accent instead of a flat grey."""
    a = [int(hex_a[i : i + 2], 16) for i in (0, 2, 4)]
    b = [int(hex_b[i : i + 2], 16) for i in (0, 2, 4)]
    return "".join(f"{round(x + (y - x) * t):02X}" for x, y in zip(a, b))


def _header(ws, headers: list[str], widths: list[int], freeze: str = "A2", accent: str | None = None) -> None:
    font = Font(name="Calibri", bold=True, color=accent or "4FFFB0", size=10)
    border = Border(bottom=Side(style="thin", color=accent or "1A2535"))
    for col, (label, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill, cell.font, cell.border = HEAD_FILL, font, border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = freeze


def _body(ws, rows: list[list], wrap_col: int | None = None, accent: str | None = None) -> None:
    stripe = PatternFill("solid", fgColor=_mix(accent, "0C1117", 0.8)) if accent else STRIPE
    for r, values in enumerate(rows, 2):
        for c, value in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=safe(value))
            cell.font, cell.border = BODY_FONT, RULE
            cell.alignment = Alignment(vertical="center", wrap_text=(c == wrap_col))
            if r % 2 == 0:
                cell.fill = stripe


def _totals(ws, rows: list[list], numeric_cols: set[int], label: str = "Total") -> None:
    """A bold sum row right under the data — only for columns it's honest to
    add up. A rate or a distinct-count column is left blank rather than
    showing a number that means nothing summed."""
    if not rows:
        return
    row_no = len(rows) + 2
    width = len(rows[0])
    top = Border(top=Side(style="thin", color="3A4A5E"))
    for c in range(1, width + 1):
        if c == 1:
            value = label
        elif c in numeric_cols:
            value = sum(row[c - 1] for row in rows if isinstance(row[c - 1], (int, float)))
        else:
            value = ""
        cell = ws.cell(row=row_no, column=c, value=value)
        cell.font, cell.border = BOLD_FONT, top


def _save(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _hours(minutes) -> float:
    """Minutes read fine on screen next to a chart; alone in a spreadsheet
    cell they don't — hours are what a person actually estimates in."""
    return round((minutes or 0) / 60, 1)


PCT_KEYS = {"completion_rate", "report_rate", "close_rate", "sla_rate", "coverage"}


def _pct(value) -> str:
    return "—" if value is None else f"{value * 100:.1f}%"


def _fmt(value):
    """None reads as a blank cell that looks broken; an em dash reads as
    'measured, and there's nothing there'."""
    return "—" if value is None else value


def _metric_rows(d: dict) -> list[list]:
    """Metric/Value pairs for a flat stats dict, skipping the nested
    breakdowns (dicts/lists) a summary sheet isn't the place for, and
    rendering known rate fields as a percentage rather than a bare 0.667."""
    return [
        [k.replace("_", " ").title(), _pct(v) if k in PCT_KEYS else _fmt(v)]
        for k, v in d.items()
        if not isinstance(v, (dict, list))
    ]


# ── work log ──────────────────────────────────────────────────────────────────

WORK_LOG_HEADERS = [
    "Date",
    "Kind",
    "Member",
    "Task Type",
    "Question Type",
    "Customer",
    "Count",
    "Effort (min)",
    "Status",
    "Due",
    "Jira",
    "Notes",
]


async def _work_log_rows(db: AsyncSession, **filters) -> tuple[list[list], bool]:
    entries, total = await entries_svc.list_entries(
        db, page=1, page_size=MAX_ROWS, **filters
    )
    rows = []
    for e in entries:
        if not e.items:
            rows.append(
                [
                    e.entry_date.isoformat(),
                    e.kind.title(),
                    e.member.display_name,
                    "",
                    "",
                    "",
                    "",
                    "",
                    STATUS_LABEL.get(e.status, e.status),
                    "",
                    "",
                    e.raw_text or "",
                ]
            )
        rows.extend(
            [
                e.entry_date.isoformat(),
                e.kind.title(),
                e.member.display_name,
                it.task_type.name,
                ", ".join(q.name for q in it.question_types),
                it.customer or "",
                it.count or "",
                it.effort_minutes or "",
                STATUS_LABEL.get(it.status, it.status),
                it.due_at.isoformat() if it.due_at else "",
                it.jira_issue_key or "",
                it.notes or e.raw_text or "",
            ]
            for it in e.items
        )
    return rows, total > len(entries)


async def work_log_xlsx(db: AsyncSession, **filters) -> bytes:
    rows, truncated = await _work_log_rows(db, **filters)
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Log"
    _header(ws, WORK_LOG_HEADERS, [12, 8, 20, 28, 16, 18, 7, 11, 12, 12, 14, 50])
    _body(ws, rows, wrap_col=12)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(WORK_LOG_HEADERS))}1"
    if truncated:
        ws.cell(
            row=len(rows) + 3,
            column=1,
            value=f"Truncated at {MAX_ROWS} entries — narrow the date range.",
        ).font = BOLD_FONT
    return _save(wb)


async def work_log_csv(db: AsyncSession, **filters) -> str:
    rows, truncated = await _work_log_rows(db, **filters)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(WORK_LOG_HEADERS)
    w.writerows([[safe(v) for v in row] for row in rows])
    if truncated:
        w.writerow([f"Truncated at {MAX_ROWS} entries"])
    return buf.getvalue()


# ── content requests ──────────────────────────────────────────────────────────


async def content_requests_xlsx(db: AsyncSession, **filters) -> bytes:
    data = await cr_svc.query(db, page=1, page_size=MAX_ROWS, **filters)
    wb = Workbook()
    ws = wb.active
    ws.title = "Content Requests"
    headers = [
        "Key",
        "Summary",
        "Status",
        "Assignee",
        "Reporter",
        "Priority",
        "Type",
        "Created",
        "Updated",
        "Due",
        "Resolved",
    ]
    _header(ws, headers, [12, 60, 18, 22, 22, 12, 18, 12, 12, 12, 12])
    _body(
        ws,
        [
            [
                r["issue_key"],
                r["summary"],
                r["status"],
                r["assignee"],
                r["reporter"] or "",
                r["priority"],
                r["issue_type"] or "",
                *(
                    v.date().isoformat() if v else ""
                    for v in (r["created_at"], r["updated_at"])
                ),
                r["due_date"].isoformat() if r["due_date"] else "",
                r["resolved_at"].date().isoformat() if r["resolved_at"] else "",
            ]
            for r in data["items"]
        ],
        wrap_col=2,
    )
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    return _save(wb)


# ── summary of the numbers behind the dashboard ───────────────────────────────


async def analytics_xlsx(db: AsyncSession, scope) -> bytes:
    """One sheet per breakdown, so the KPI cards can be checked against rows."""
    from services import analytics as an

    wb = Workbook()
    wb.remove(wb.active)

    async def sheet(title: str, headers: list[str], rows: list[list], widths=None):
        ws = wb.create_sheet(title[:31])
        _header(ws, headers, widths or [22] * len(headers))
        _body(ws, rows)

    s = await an.summary(db, scope)
    await sheet(
        "Summary",
        ["Metric", "Value"],
        [
            [k.replace("_", " ").title(), v]
            for k, v in s.items()
            if not isinstance(v, dict)
        ],
        [28, 16],
    )
    await sheet(
        "By member",
        [
            "Member",
            "Tasks",
            "Items",
            "Effort (min)",
            *[x.replace("_", " ").title() for x in STATUSES],
            "Completion",
        ],
        [
            [
                r["member"],
                r["tasks"],
                r["volume"],
                r["effort_minutes"],
                *[r[x] for x in STATUSES],
                r["completion_rate"],
            ]
            for r in await an.by_member(db, scope)
        ],
    )
    await sheet(
        "By task type",
        [
            "Task type",
            "Tasks",
            "Items",
            "Effort (min)",
            *[x.replace("_", " ").title() for x in STATUSES],
        ],
        [
            [
                r["task_type"],
                r["tasks"],
                r["volume"],
                r["effort_minutes"],
                *[r[x] for x in STATUSES],
            ]
            for r in await an.by_task_type(db, scope)
        ],
        [32, 10, 10, 12, 10, 12, 10, 10],
    )
    await sheet(
        "Plan adherence",
        [
            "Member",
            "Planned",
            "Reported",
            "Closed",
            "No update",
            "Report rate",
            "Close rate",
        ],
        [
            [
                r["member"],
                r["planned"],
                r["reported"],
                r["closed"],
                r["no_update"],
                r["report_rate"],
                r["close_rate"],
            ]
            for r in await an.plan_adherence(db, scope)
        ],
    )
    await sheet(
        "By customer",
        ["Customer", "Tasks", "Items", "Effort (min)", "Outstanding"],
        [
            [
                r["customer"],
                r["tasks"],
                r["volume"],
                r["effort_minutes"],
                r["outstanding"],
            ]
            for r in await an.by_customer(db, scope, limit=100)
        ],
        [32, 10, 10, 12, 14],
    )
    return _save(wb)


# ── team overview (Overview tab) ────────────────────────────────────────────


def _sheet_builder(wb: Workbook):
    """Each call gets the next accent in the palette — a distinct header
    colour, stripe tint and tab colour per sheet, and a bold sum row for
    whichever columns are honest to add up."""
    palette = iter(ACCENTS * 3)

    async def sheet(title: str, headers: list[str], rows: list[list], widths=None, totals: set[int] | None = None):
        accent = next(palette)
        ws = wb.create_sheet(title[:31])
        ws.sheet_properties.tabColor = accent
        _header(ws, headers, widths or [22] * len(headers), accent=accent)
        _body(ws, rows, accent=accent)
        if rows and len(headers) > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        if totals:
            _totals(ws, rows, totals)
        return ws

    return sheet


async def team_overview_xlsx(db: AsyncSession, scope) -> bytes:
    """Everything the Overview screen shows, one sheet per card, for the
    period the page's own date range is set to. No ticket keys, no
    hyperlinks — those belong to the work-log export, not this one."""
    from services import analytics as an

    wb = Workbook()
    wb.remove(wb.active)
    sheet = _sheet_builder(wb)

    await sheet("Summary", ["Metric", "Value"], _metric_rows(await an.summary(db, scope)), [28, 16])

    await sheet(
        "Trend",
        ["Date", "Tasks", "Items", "Effort (hrs)", "Closed", "Plans", "Updates"],
        [
            [r["date"], r["tasks"], r["volume"], _hours(r["effort_minutes"]), r["closed"], r["plans"], r["updates"]]
            for r in await an.trend(db, scope)
        ],
        [14, 10, 10, 14, 10, 10, 10],
        totals={2, 3, 4, 5, 6, 7},
    )

    status_headers = [x.replace("_", " ").title() for x in STATUSES]
    await sheet(
        "By member",
        ["Member", "Tasks", "Items", "Effort (hrs)", *status_headers, "Completion"],
        [
            [r["member"], r["tasks"], r["volume"], _hours(r["effort_minutes"]), *[r[x] for x in STATUSES], _pct(r["completion_rate"])]
            for r in await an.by_member(db, scope)
        ],
        totals={2, 3, 4, 5, 6, 7, 8},
    )

    await sheet(
        "By area",
        ["Area", "Tasks", "Items", "Effort (hrs)", "Members", "Customers", *status_headers],
        [
            [r["label"], r["tasks"], r["volume"], _hours(r["effort_minutes"]), r["members"], r["customers"], *[r[x] for x in STATUSES]]
            for r in await an.by_area(db, scope)
        ],
        [26, 10, 10, 14, 10, 12, 10, 12, 10, 10],
        totals={2, 3, 4, 7, 8, 9, 10},
    )

    await sheet(
        "By task type",
        ["Task type", "Tasks", "Items", "Effort (hrs)", *status_headers],
        [
            [r["task_type"], r["tasks"], r["volume"], _hours(r["effort_minutes"]), *[r[x] for x in STATUSES]]
            for r in await an.by_task_type(db, scope)
        ],
        [32, 10, 10, 14, 10, 12, 10, 10],
        totals={2, 3, 4, 5, 6, 7, 8},
    )

    due = await an.due_risk(db, scope, today())
    await sheet("Due risk", ["Bucket", "Count"], _metric_rows(due), [24, 12])

    await sheet("Cycle time", ["Metric", "Value"], _metric_rows(await an.cycle_time(db, scope)), [24, 14])

    return _save(wb)


# ── per-member overview (Member Detail tab) ─────────────────────────────────


async def member_overview_xlsx(db: AsyncSession, member, scope) -> bytes:
    """Everything the Member Detail screen shows for one person, for the
    period the page's own date range is set to. No entries table, no ticket
    keys — this is the breakdown, not the log."""
    from services import analytics as an

    wb = Workbook()
    wb.remove(wb.active)
    sheet = _sheet_builder(wb)

    profile_rows = [
        ["Name", member.display_name],
        ["Role", member.role.title()],
        ["Email", member.email or "—"],
        ["From", scope.frm.isoformat()],
        ["To", scope.to.isoformat()],
        *_metric_rows(await an.summary(db, scope)),
    ]
    await sheet("Profile", ["Metric", "Value"], profile_rows, [24, 24])

    status_headers = [x.replace("_", " ").title() for x in STATUSES]
    await sheet(
        "By stream",
        ["Stream", "Tasks", "Items", "Effort (hrs)", *status_headers],
        [
            [r["label"], r["tasks"], r["volume"], _hours(r["effort_minutes"]), *[r[x] for x in STATUSES]]
            for r in await an.by_pipeline(db, scope)
        ],
        [28, 10, 10, 14, 10, 12, 10, 10],
        totals={2, 3, 4, 5, 6, 7, 8},
    )

    await sheet(
        "Work areas",
        ["Task type", "Tasks", "Items", "Effort (hrs)", *status_headers],
        [
            [r["task_type"], r["tasks"], r["volume"], _hours(r["effort_minutes"]), *[r[x] for x in STATUSES]]
            for r in await an.by_task_type(db, scope)
        ],
        [32, 10, 10, 14, 10, 12, 10, 10],
        totals={2, 3, 4, 5, 6, 7, 8},
    )

    await sheet(
        "Customers",
        ["Customer", "Tasks", "Items", "Effort (hrs)", "Outstanding"],
        [
            [r["customer"], r["tasks"], r["volume"], _hours(r["effort_minutes"]), r["outstanding"]]
            for r in await an.by_customer(db, scope, limit=100)
        ],
        [32, 10, 10, 14, 14],
        totals={2, 3, 4, 5},
    )

    await sheet(
        "Question types",
        ["Question type", "Tasks", "Items"],
        [[r["question_type"], r["tasks"], r["volume"]] for r in await an.by_question_type(db, scope)],
        [32, 10, 10],
        totals={2, 3},
    )

    await sheet(
        "Plan adherence",
        ["Planned", "Reported", "Closed", "No update", "Report rate", "Close rate"],
        [
            [r["planned"], r["reported"], r["closed"], r["no_update"], _pct(r["report_rate"]), _pct(r["close_rate"])]
            for r in await an.plan_adherence(db, scope)
        ],
        [12, 12, 12, 12, 14, 12],
    )

    await sheet(
        "Output over time",
        ["Date", "Tasks", "Items", "Effort (hrs)", "Closed", "Plans", "Updates"],
        [
            [r["date"], r["tasks"], r["volume"], _hours(r["effort_minutes"]), r["closed"], r["plans"], r["updates"]]
            for r in await an.trend(db, scope)
        ],
        [14, 10, 10, 14, 10, 10, 10],
        totals={2, 3, 4, 5, 6, 7},
    )

    await sheet("Cycle time", ["Metric", "Value"], _metric_rows(await an.cycle_time(db, scope)), [24, 14])

    return _save(wb)
