"""Exports. The formula guard matters most — notes are free text that lands in
a spreadsheet cell, and Excel runs anything starting with = + - @."""

import csv
import io

import pytest
from openpyxl import load_workbook

from services.export import safe

DAY = "2030-09-10"
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.mark.parametrize("raw", [
    '=HYPERLINK("http://evil","claim")',
    "+1234",
    "-cmd|' /c calc'!A0",
    "@SUM(A1:A9)",
    "\tstartswith tab",
])
def test_dangerous_cells_are_quoted(raw):
    assert safe(raw) == "'" + raw


@pytest.mark.parametrize("raw", ["Content review", "3 items", "", "a-b", None, 42])
def test_ordinary_values_pass_through(raw):
    assert safe(raw) == raw


@pytest.fixture
async def logged(client, member, task_type):
    plan = (await client.post("/api/entries/plans", json={
        "member_id": member, "entry_date": DAY,
        "items": [{"task_type_id": task_type, "count": 4, "customer": "Acme",
                   "notes": '=cmd|calc'}],
    })).json()
    return plan


async def test_work_log_xlsx_is_a_real_workbook(client, logged, member):
    r = await client.get("/api/exports/work-log.xlsx",
                         params={"from": DAY, "to": DAY, "member_id": member})
    assert r.status_code == 200 and r.headers["content-type"] == XLSX
    assert "work-log-2030-09-10" in r.headers["content-disposition"]

    ws = load_workbook(io.BytesIO(r.content)).active
    header = [c.value for c in ws[1]]
    assert header[:4] == ["Date", "Kind", "Member", "Task Type"]
    row = [c.value for c in ws[2]]
    assert row[0] == DAY and row[6] == 4 and row[5] == "Acme"
    assert row[10] == "'=cmd|calc", "notes must be neutralised in the cell"


async def test_work_log_csv_quotes_formulas_too(client, logged, member):
    r = await client.get("/api/exports/work-log.csv",
                         params={"from": DAY, "to": DAY, "member_id": member})
    assert r.status_code == 200
    rows = list(csv.reader(io.StringIO(r.text)))
    assert rows[0][0] == "Date"
    assert rows[1][10] == "'=cmd|calc"


async def test_ae_daily_xlsx_lays_out_dates_over_members(client, ae_member):
    await client.put("/api/ae/daily", json={
        "member_id": ae_member, "entry_date": DAY, "notes": "steady",
        "metrics": {"bug_fixes": 7},
    })
    r = await client.get("/api/exports/ae-daily.xlsx", params={"from": DAY, "to": DAY})
    assert r.status_code == 200

    ws = load_workbook(io.BytesIO(r.content)).active
    assert ws.cell(row=1, column=1).value == "Metric"
    assert ws.cell(row=1, column=2).value == DAY        # date band
    assert ws.cell(row=2, column=2).value == "PyTest AE"  # member under it
    labels = [ws.cell(row=r_, column=1).value for r_ in range(3, 14)]
    assert "Bug Fixes" in labels and labels[-1] == "Notes"
    assert ws.cell(row=3 + labels.index("Bug Fixes"), column=2).value == 7


async def test_analytics_workbook_has_a_sheet_per_breakdown(client, logged, member):
    r = await client.get("/api/exports/analytics.xlsx",
                         params={"from": DAY, "to": DAY, "member_id": member})
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Summary", "By member", "By task type",
                             "Plan adherence", "By customer"]
    assert wb["By member"].cell(row=2, column=1).value == "PyTest Member"


async def test_content_requests_export_is_empty_but_valid(client):
    r = await client.get("/api/exports/content-requests.xlsx")
    assert r.status_code == 200
    ws = load_workbook(io.BytesIO(r.content)).active
    assert [c.value for c in ws[1]][:3] == ["Key", "Summary", "Status"]
